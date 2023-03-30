from django.contrib import messages
import openpyxl
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.db.models import Q
from django.shortcuts import redirect, get_object_or_404, render
from django.urls import reverse_lazy
from .filters import CommentFilter, LeadContactFilter
from django.views import View
from django_filters.views import FilterView
from django.views.generic import ListView, DetailView, DeleteView, UpdateView, CreateView
from .models import Lead, Team, LeadFilter, User, Comment, Contact
from client.models import Client, Comment as ClientComment, Contact as ClientContact
from .forms import AddCommentForm, AddFileForm, AddContactForm


class LeadListView(LoginRequiredMixin,FilterView):
    paginate_by = 10
    model = Lead
    template_name = 'lead/lead_list.html'
    context_object_name='leads'
    filter_class = LeadFilter
    def get_queryset(self):
        team = Team.objects.filter(members__id=self.request.user.id)[0]
        queryset = super(LeadListView, self).get_queryset().order_by('country')
        if self.request.user.is_superuser or team.name == 'Operation':
            return queryset.filter(converted_to_client = False)
        elif self.request.user.groups.all()[0].name=='teamlead':
            return queryset.filter(team = team, converted_to_client = False)
        else:
            return queryset.filter(assign_to = self.request.user, converted_to_client = False)
    
class LeadDetailView(LoginRequiredMixin,DetailView): 
    model = Lead
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['commentform'] = AddCommentForm()
        context['fileform'] = AddFileForm()
        context['title'] = 'View Lead Details'
        return context
        
    def get_queryset(self):
        queryset = super(LeadDetailView, self).get_queryset()
        return queryset.filter(pk=self.kwargs.get('pk'))
        
class LeadUpdateView(LoginRequiredMixin,UpdateView):
    model = Lead    
    fields = ('contact_name','company_name','address','country','region','phone','email','profile','care_update','portfolio','source','priority','status','team','assign_to','created_by',)
    success_url = reverse_lazy('leads:list')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Update Lead'
        context['commentform'] = AddCommentForm()
        context['fileform'] = AddFileForm()

        return context
    
    def get_queryset(self):
        queryset = super(LeadUpdateView, self).get_queryset()
        return queryset.filter(pk=self.kwargs.get('pk'))
        
class LeadCreateView(LoginRequiredMixin,CreateView):
    model = Lead    
    fields = ('contact_name','company_name','address','country','region','phone','email','profile','care_update','portfolio','source','priority','status','team','assign_to','created_by',)
    success_url = reverse_lazy('leads:list',)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Add lead'

        return context
 
    def form_valid(self, form):
        self.object = form.save(commit=False)
        self.object.save()
        
        return redirect(self.get_success_url())
    
class LeadDeleteView(LoginRequiredMixin, DeleteView):
    model = Lead
    success_url = reverse_lazy('leads:list')

    def get_queryset(self):
        queryset = super(LeadDeleteView, self).get_queryset()
        if self.request.user.username == 'haict':
            return queryset.filter(pk=self.kwargs.get('pk'))
        else:
            return queryset.filter(assign_to=self.request.user, pk=self.kwargs.get('pk'))
        
    def get(self, request, *args, **kwargs):
        return self.post(request, *args, **kwargs)

class AddContactView(LoginRequiredMixin,CreateView):
    model = Contact
    fields = ['first_name','last_name','lead',
                  'birth_date','married','family','phone_number','email','religion',
                  'disc','gains','pains','stakeholders',
                  'assign_to','team']
     
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)       
        context['title'] = 'Add Contact'
        context['form'] = AddContactForm()
        context['lead'] = get_object_or_404(Lead,pk=self.kwargs['pk'])

        return context

    def form_valid(self, form):
        print('Here')
        lead = Lead.objects.filter(pk=self.kwargs['pk'])[0]
        self.object = form.save(commit=False)
        self.object.lead = lead
        self.object.save() 
        return redirect('leads:contact')

class AddCommentView(LoginRequiredMixin,View):
    def post(self,request,*args, **kwargs):
        pk= self.kwargs.get('pk')
        form=AddCommentForm(request.POST)
        
        if form.is_valid():
            team=Team.objects.filter(members__id=self.request.user.id)[0]
            lead = Lead.objects.get(pk=pk)
            comment=form.save(commit=False)
            comment.team = team
            comment.created_by = request.user
            comment.lead_id=pk
            comment.save()
            lead.care_update = comment.content
            lead.save()
            
        return redirect('leads:detail',pk=pk)

class AddFileView(LoginRequiredMixin,View):
    
    def post(self,request,*args, **kwargs):
        pk= self.kwargs.get('pk')
        form=AddFileForm(request.POST,request.FILES)
        
        if form.is_valid():
            team=Team.objects.filter(members__id=self.request.user.id)[0]
            file=form.save(commit=False)
            file.team = team
            file.created_by = request.user
            file.lead_id=pk
            file.save()
            
        return redirect('leads:detail',pk=pk)


class ConvertView(View):
    def get(self, request, *args, **kwargs):
        pk=self.kwargs.get('pk')
        if self.request.user.username == 'haict':            
            lead = get_object_or_404(Lead,pk=pk)
        else:
            lead = get_object_or_404(Lead, assign_to=self.request.user, pk=pk)
        team = Team.objects.filter(members__id=request.user.id)[0]
        
        client = Client.objects.create(
            team = team,
            contact_name = lead.contact_name,
            company_name = lead.company_name,
            address = lead.address,
            country = lead.country,
            region = lead.region,
            phone = lead.phone,
            email = lead.email,
            profile = lead.profile,
            care_update = lead.care_update,
            portfolio = lead.portfolio,
            source = lead.source,
            assign_to = lead.assign_to,
            created_by = request.user,
        ) 
        lead.converted_to_client = True
        lead.status = '5.ordered'
        lead.save()
        
        #convert lead cmt to client cmt
        comments = lead.comments.all()
        for comment in comments:
            ClientComment.objects.create(
                client = client,
                content = comment.content,
                created_by = comment.created_by,
                team = team                
            )

        #convert lead contact to client contact
        contacts = lead.contacts.all()
        for contact in contacts:
            ClientContact.objects.create(
                client = client,
                first_name = contact.first_name,
                last_name = contact.last_name,
                birth_date = contact.birth_date,
                married = contact.married,
                family = contact.family,
                phone_number = contact.phone_number,
                email = contact.email,
                religion = contact.religion,
                disc = contact.disc,
                stakeholders = contact.stakeholders,
                gains = contact.gains,
                pains = contact.pains,
                assign_to = contact.assign_to,
                team = lead.team                
            )
            contact.delete()

        messages.success(request,'The Lead Has Been Converted!')
        return redirect('leads:list')
    
class SearchLead(LoginRequiredMixin, ListView):
    model = Lead
    template_name = "lead/search_lead.html"

    def get_queryset(self):  # new
        query = self.request.GET.get("q")
        object_list = Lead.objects.filter(
            Q(company_name__icontains=query) | Q(contact_name__icontains=query)
        )
        return object_list

class CommentList(LoginRequiredMixin, FilterView):
    paginate_by = 10
    template_name = 'lead/comment.html'
    context_object_name='comments'
    filterset_class = CommentFilter
         
    def get_queryset(self):
        comments = Comment.objects.all()
        if self.request.user.is_superuser:
            return comments.order_by('-created_at')
        else:
            return comments.filter(created_by=self.request.user).order_by('-created_at') 

class ContactList(LoginRequiredMixin,FilterView):
    paginate_by = 5
    context_object_name = 'contacts'
    template_name = 'lead/contact.html'
    filterset_class = LeadContactFilter

    def get_queryset(self):
        contacts = Contact.objects.all()
        if self.request.user.is_superuser:
            return contacts.order_by('lead')
        else:
            return contacts.filter(assign_to=self.request.user).order_by('lead')

class ContactDetailView(LoginRequiredMixin,DetailView): 
    model = Contact
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Contact Detail'
        return context
        
    def get_queryset(self):
        queryset = super(ContactDetailView, self).get_queryset()
        return queryset.filter(pk=self.kwargs.get('pk'))

class ContactUpdateView(LoginRequiredMixin,UpdateView):
    model = Contact    
    fields = ['first_name','last_name','lead',
                  'birth_date','married','family','phone_number','email','religion',
                  'disc','gains','pains','stakeholders',
                  'assign_to','team']
    success_url = reverse_lazy('leads:contact')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Update Contact'
        return context
    
    def get_queryset(self):
        queryset = super(ContactUpdateView, self).get_queryset()
        return queryset.filter(pk=self.kwargs.get('pk'))
    
@login_required
def delete_contact(request, pk):    
    contact = Contact.objects.get(id=pk)
    contact.delete()
    messages.success(
            request, 'Task have been deleted successfully'
            )
    return redirect('leads:contact')
    
@login_required 
def importLead(request):
    if request.method == 'POST':
        if 'myfile' in request.FILES:
            excel_file = request.FILES['myfile']
            wb = openpyxl.load_workbook(excel_file)
            worksheet = wb['Sheet1']
            data = list()
            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                lead = Lead()
                team=Team.objects.filter(name=row[0].value)[0]
                lead.team = team
                lead.contact_name = row[1].value
                lead.company_name = row[2].value
                lead.address = row[3].value
                lead.country = row[4].value
                lead.region = row[5].value
                lead.phone = row[6].value
                lead.email = row[7].value
                lead.profile = row[8].value
                lead.care_update = row[9].value
                lead.portfolio = row[10].value
                lead.priority = row[11].value
                lead.status = row[12].value
                lead.source = row[13].value
                pic=User.objects.filter(username=row[14].value)[0]
                lead.assign_to = pic
                lead.created_by = request.user
                data.append(lead)
                
            Lead.objects.bulk_create(data)
            return redirect('leads:list')
        else:
            messages.warning(request, message='Failed to upload data!')
            return render(request, 'lead/lead_import.html')
    else:
        return render(request, 'lead/lead_import.html')